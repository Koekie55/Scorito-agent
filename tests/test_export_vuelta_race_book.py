from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from scripts.export_vuelta_race_book import (
    _future_stages,
    _race_state,
    export_workbook,
)


def _write(path: Path, content: object) -> None:
    path.write_text(json.dumps({"Content": content}), encoding="utf-8")


def test_race_state_uses_first_stage_without_results(tmp_path: Path) -> None:
    _write(
        tmp_path / "marketroundstage.json",
        [
            {"StageId": 102, "StageOrder": 2},
            {"StageId": 101, "StageOrder": 1},
            {"StageId": 103, "StageOrder": 3},
        ],
    )
    _write(tmp_path / "stageresult_rider_101.json", [{"RiderId": 1}])
    _write(tmp_path / "stageresult_rider_102.json", [{"RiderId": 2}])
    _write(tmp_path / "stageresult_rider_103.json", [])

    completed, next_stage = _race_state(tmp_path)

    assert [stage["stage_no"] for stage in completed] == [1, 2]
    assert next_stage == 3


def test_future_stages_begin_with_next_uncompleted_stage() -> None:
    personal = {"stages": [{"stage_no": number} for number in range(1, 22)]}

    assert [stage["stage_no"] for stage in _future_stages(personal, 3)] == list(
        range(3, 22)
    )
    assert _future_stages(personal, None) == []


def test_export_workbook_writes_valid_actionable_race_book(tmp_path: Path) -> None:
    output = export_workbook(tmp_path / "race-book.xlsx")

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "Dashboard",
        "Current Squad",
        "Forward Stage Plan",
        "Stage Matrix",
        "Completed Stages",
        "Model Comparator",
        "Sources & Validation",
    ]
    assert workbook["Current Squad"].max_row == 24
    forward_plan = workbook["Forward Stage Plan"]
    assert forward_plan.auto_filter.ref == f"A4:S{forward_plan.max_row}"
    workbook.close()