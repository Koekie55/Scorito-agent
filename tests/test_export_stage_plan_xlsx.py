from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from scripts.export_stage_plan_xlsx import _current_forward_plan, export_workbook


def _write(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload), encoding="utf-8")


def test_current_forward_plan_uses_saved_result_boundary(tmp_path: Path) -> None:
    plan = {
        "completed_stages_excluded_from_forward_plan": [1],
        "forward_stage_plan": [{"stage_no": stage_no} for stage_no in range(2, 5)],
    }
    _write(
        tmp_path / "marketroundstage.json",
        {"Content": [
            {"StageId": 101, "StageOrder": 1},
            {"StageId": 102, "StageOrder": 2},
            {"StageId": 103, "StageOrder": 3},
            {"StageId": 104, "StageOrder": 4},
        ]},
    )
    _write(tmp_path / "stageresult_rider_101.json", {"Content": [{"RiderId": 1}]})
    _write(tmp_path / "stageresult_rider_102.json", {"Content": [{"RiderId": 2}]})
    _write(tmp_path / "stageresult_rider_103.json", {"Content": []})

    completed, forward = _current_forward_plan(plan, tmp_path)

    assert completed == [1, 2]
    assert [stage["stage_no"] for stage in forward] == [3, 4]


def test_export_uses_current_forward_stages_and_totals(tmp_path: Path) -> None:
    source_dir = Path(__file__).resolve().parents[1] / "data" / "scorito" / "vuelta2026"
    for name in (
        "hawktuah_candidate_stage_plan.json",
        "stage_top20_predictions.json",
        "marketroundstage.json",
        "stageresult_rider_2820.json",
        "stageresult_rider_2821.json",
        "stageresult_rider_2822.json",
    ):
        (tmp_path / name).write_bytes((source_dir / name).read_bytes())
    output = tmp_path / "stage-plan.xlsx"

    export_workbook(
        tmp_path / "hawktuah_candidate_stage_plan.json",
        tmp_path / "stage_top20_predictions.json",
        output,
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    assert workbook["Stage Plan"].max_row == 20
    assert workbook["Stage Plan"]["A2"].value == 3
    summary = dict(workbook["Summary"].iter_rows(min_row=2, values_only=True))
    assert summary["Forward enrolled-stage points"] == 6113
    assert summary["Final forward projected score"] == 6400.93
    assert summary["Important"].startswith("Completed stages excluded: 1, 2.")
    workbook.close()