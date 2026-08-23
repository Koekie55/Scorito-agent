"""Export a saved Scorito stage plan JSON to a practical Excel workbook."""
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PLAN = ROOT / "data" / "scorito" / "vuelta2026" / "hawktuah_candidate_stage_plan.json"
DEFAULT_PREDICTIONS = ROOT / "data" / "scorito" / "vuelta2026" / "stage_top20_predictions.json"
STAGE_POINTS = {
    1: 50, 2: 44, 3: 40, 4: 36, 5: 32, 6: 30, 7: 28, 8: 26, 9: 24,
    10: 22, 11: 20, 12: 18, 13: 16, 14: 14, 15: 12, 16: 10,
    17: 8, 18: 6, 19: 4, 20: 1,
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
CAPTAIN_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")


def _content(path: Path) -> Any:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    return payload.get("Content", payload) if isinstance(payload, dict) else payload


def _current_forward_plan(
    plan: dict[str, Any], data_dir: Path
) -> tuple[list[int], list[dict[str, Any]]]:
    rounds_path = data_dir / "marketroundstage.json"
    completed = [int(stage_no) for stage_no in plan.get("completed_stages_excluded_from_forward_plan", [])]
    if rounds_path.exists():
        completed = []
        rounds = sorted(_content(rounds_path), key=lambda row: int(row["StageOrder"]))
        for stage in rounds:
            result_path = data_dir / f"stageresult_rider_{int(stage['StageId'])}.json"
            results = _content(result_path) if result_path.exists() else []
            if not results:
                break
            completed.append(int(stage["StageOrder"]))
    completed_set = set(completed)
    forward = [
        stage
        for stage in plan["forward_stage_plan"]
        if int(stage["stage_no"]) not in completed_set
    ]
    return completed, forward


def _name_key(value: str) -> tuple[str, ...]:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_name = "".join(char for char in normalized if not unicodedata.combining(char))
    return tuple(sorted(re.findall(r"[a-z0-9]+", ascii_name.lower())))


def _style_table(sheet, *, freeze: str = "A2") -> None:
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for column in sheet.columns:
        width = min(45, max(len(str(cell.value or "")) for cell in column) + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def _stage_reserves(plan: dict[str, Any], predictions: dict[str, Any]) -> dict[int, str]:
    display = {_name_key(name): name for name in plan["squad"]}
    selected_by_stage = {
        int(row["stage_no"]): {_name_key(name) for name in row["lineup"]}
        for row in plan["forward_stage_plan"]
    }
    reserves: dict[int, str] = {}
    for stage in predictions["stages"]:
        stage_no = int(stage["stage_no"])
        if stage_no not in selected_by_stage:
            continue
        ranks = {_name_key(row["rider"]): int(row["predicted_finish"]) for row in stage["top_20"]}
        candidates = sorted(
            display,
            key=lambda key: (-STAGE_POINTS.get(ranks.get(key, 0), 0), ranks.get(key, 999), display[key]),
        )
        reserves[stage_no] = display[next(key for key in candidates if key not in selected_by_stage[stage_no])]
    return reserves


def export_workbook(plan_path: Path, predictions_path: Path, output_path: Path) -> None:
    plan = json.loads(plan_path.read_text(encoding="utf-8"))
    predictions = json.loads(predictions_path.read_text(encoding="utf-8"))
    completed_stages, forward_stage_plan = _current_forward_plan(plan, plan_path.parent)
    current_plan = {**plan, "forward_stage_plan": forward_stage_plan}
    display = {_name_key(name): name for name in plan["squad"]}
    reserves = _stage_reserves(current_plan, predictions)
    forward_stage_points = sum(
        float(stage["projected_stage_points"]) for stage in forward_stage_plan
    )
    final_forward_score = forward_stage_points + float(
        plan["projected_classification_jersey_points"]
    )
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Field", "Value"])
    member = plan["member_claim"]
    legality = plan["legality"]
    for row in [
        ("Owner claim", member["owner"]), ("Team", member["team"]),
        ("Subleague", member["subleague"]),
        ("Ownership verification", plan["ownership_verification"]["status"]),
        ("Market snapshot UTC", plan["market_snapshot_time"]),
        ("Prediction snapshot UTC", plan["prediction_generated_at"]),
        ("Squad price", legality["current_price"]), ("Budget", plan["market_budget"]),
        ("Budget remaining", legality["budget_remaining"]),
        ("Credible sprint options", legality["credible_sprint_option_count"]),
        ("Forward enrolled-stage points", forward_stage_points),
        ("Classification/jersey points", plan["projected_classification_jersey_points"]),
        ("Final forward projected score", final_forward_score),
        (
            "Important",
            f"Completed stages excluded: {', '.join(map(str, completed_stages)) or 'none'}. "
            "Do not submit automatically.",
        ),
    ]:
        summary.append(row)
    _style_table(summary)
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 90

    stage_sheet = workbook.create_sheet("Stage Plan")
    stage_sheet.append(["Stage", "Profile", "Captain", "Projected points", "First reserve", *[f"Rider {i}" for i in range(1, 10)]])
    for stage in forward_stage_plan:
        lineup = [display[_name_key(name)] for name in stage["lineup"]]
        captain = display[_name_key(stage["captain"])]
        stage_sheet.append([stage["stage_no"], stage["profile_type"], captain, stage["projected_stage_points"], reserves[int(stage["stage_no"])], *lineup])
        for cell in stage_sheet[stage_sheet.max_row][5:]:
            if cell.value == captain:
                cell.fill = CAPTAIN_FILL
                cell.font = Font(bold=True)
    _style_table(stage_sheet)

    starts = Counter()
    captaincies = Counter()
    for stage in forward_stage_plan:
        starts.update(display[_name_key(name)] for name in stage["lineup"])
        captaincies[display[_name_key(stage["captain"])]] += 1
    sprint_options = set(legality["credible_sprint_options"])
    squad_sheet = workbook.create_sheet("Squad")
    squad_sheet.append(["Rider", "Forward starts", "Captaincies", "Credible sprint option"])
    for rider in sorted(plan["squad"], key=lambda name: (-starts[name], name)):
        squad_sheet.append([rider, starts[rider], captaincies[rider], "Yes" if rider in sprint_options else "No"])
    _style_table(squad_sheet)

    checks_sheet = workbook.create_sheet("Validation")
    checks_sheet.append(["Check", "Passed"])
    for check, passed in plan["validation"].items():
        checks_sheet.append([check.replace("_", " ").title(), "Yes" if passed else "No"])
        if passed:
            checks_sheet.cell(checks_sheet.max_row, 2).fill = OK_FILL
    checks_sheet.append(["Ownership verified in Scorito", "No - authenticated league lookup required"])
    checks_sheet.append(["Current optimum delta", "Paret-Peintre replaces Laporte in an unlocked rebuild"])
    _style_table(checks_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    reopened = load_workbook(output_path, read_only=True, data_only=False)
    if reopened.sheetnames != ["Summary", "Stage Plan", "Squad", "Validation"]:
        raise RuntimeError("saved workbook sheet validation failed")
    if reopened["Stage Plan"].max_row != len(forward_stage_plan) + 1:
        raise RuntimeError("saved workbook stage count validation failed")
    reopened.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--plan", type=Path, default=DEFAULT_PLAN)
    parser.add_argument("--predictions", type=Path, default=DEFAULT_PREDICTIONS)
    parser.add_argument("--out", type=Path, default=DEFAULT_PLAN.with_suffix(".xlsx"))
    args = parser.parse_args()
    export_workbook(args.plan, args.predictions, args.out)
    print(args.out)


if __name__ == "__main__":
    main()
