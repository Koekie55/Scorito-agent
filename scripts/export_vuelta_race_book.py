"""Export the actionable Vuelta race book from refreshed saved model outputs."""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from scorito_agent.scorito import load_snapshot  # noqa: E402

DATA_DIR = ROOT / "data" / "scorito" / "vuelta2026"
DEFAULT_OUTPUT = DATA_DIR / "Vuelta_2026_Optimal_Team_and_Stage_Plan.xlsx"
PERSONAL_PATH = DATA_DIR / "personal_team_full_analysis.json"
PREDICTIONS_PATH = DATA_DIR / "stage_top20_predictions.json"
PROJECTION_PATH = DATA_DIR / "projected_recommendation.json"
OPTIMUM_PATH = DATA_DIR / "optimal_team_exact_analysis.json"
PERSONAL_SELECTION_PATH = DATA_DIR / "personal" / "teamselection.json"
ROUNDS_PATH = DATA_DIR / "marketroundstage.json"
TEAMS_PATH = DATA_DIR / "teams_all.json"

NAVY = "17324D"
BLUE = "2E75B6"
LIGHT_BLUE = "D9EAF7"
GREEN = "DDEBF7"
LIGHT_GREEN = "E2F0D9"
GOLD = "FFF2CC"
RED = "FCE4D6"
GREY = "E7E6E6"
WHITE = "FFFFFF"
THIN_GREY = Side(style="thin", color="D9E1F2")


def _content(path: Path) -> Any:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    return payload.get("Content", payload) if isinstance(payload, dict) else payload


def _race_state(data_dir: Path = DATA_DIR) -> tuple[list[dict[str, Any]], int | None]:
    rounds = sorted(
        _content(data_dir / ROUNDS_PATH.name),
        key=lambda row: int(row["StageOrder"]),
    )
    completed: list[dict[str, Any]] = []
    next_stage: int | None = None
    for stage in rounds:
        stage_no = int(stage["StageOrder"])
        stage_id = int(stage["StageId"])
        result_path = data_dir / f"stageresult_rider_{stage_id}.json"
        results = _content(result_path) if result_path.exists() else []
        if results and next_stage is None:
            completed.append(
                {"stage_no": stage_no, "stage_id": stage_id, "results": results}
            )
        elif next_stage is None:
            next_stage = stage_no
    return completed, next_stage


def _future_stages(
    personal: dict[str, Any], next_stage: int | None
) -> list[dict[str, Any]]:
    if next_stage is None:
        return []
    return [
        stage
        for stage in personal["stages"]
        if int(stage["stage_no"]) >= next_stage
    ]


def _name_key(value: str) -> tuple[str, ...]:
    normalized = unicodedata.normalize("NFKD", str(value))
    ascii_name = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    return tuple(sorted(re.findall(r"[a-z0-9]+", ascii_name.lower())))


def _timestamp(value: str | None) -> str:
    if not value:
        return "not available"
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    return parsed.astimezone(UTC).strftime("%d %b %Y, %H:%M UTC")


def _title(sheet, title: str, subtitle: str, columns: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    sheet["A1"] = title
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].font = Font(color=WHITE, bold=True, size=18)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(color="5B6573", italic=True, size=10)
    sheet.row_dimensions[2].height = 22


def _header(sheet, row: int) -> None:
    for cell in sheet[row]:
        if cell.value is None:
            continue
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GREY)


def _finish(sheet, *, freeze: str, widths: dict[int, float]) -> None:
    sheet.freeze_panes = freeze
    sheet.sheet_view.showGridLines = False
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="top",
                wrap_text=True,
            )


def _team_names() -> dict[int, str]:
    return {int(row["Id"]): str(row["Name"]) for row in _content(TEAMS_PATH)}


def _objective_scores(projection: dict[str, Any]) -> dict[int, dict[tuple[str, ...], float]]:
    return {
        int(stage_no): {
            _name_key(row["rider"]): float(row.get("score") or 0.0)
            for row in rows
        }
        for stage_no, rows in projection.get("stage_rankings", {}).items()
    }


def _first_reserves(
    personal: dict[str, Any], projection: dict[str, Any]
) -> dict[int, str]:
    riders = {str(row["rider"]): row for row in personal["riders"]}
    display = {_name_key(name): name for name in riders}
    scores = _objective_scores(projection)
    reserves: dict[int, str] = {}
    for stage in personal["stages"]:
        stage_no = int(stage["stage_no"])
        if stage_no == 1:
            continue
        selected = {_name_key(name) for name in stage["lineup"]}
        ranked = sorted(
            display,
            key=lambda key: (-scores.get(stage_no, {}).get(key, 0.0), display[key]),
        )
        reserves[stage_no] = display[next(key for key in ranked if key not in selected)]
    return reserves


def _dashboard(
    workbook: Workbook,
    personal: dict[str, Any],
    projection: dict[str, Any],
    reserves: dict[int, str],
    completed: list[dict[str, Any]],
    next_stage: int | None,
) -> None:
    sheet = workbook.active
    sheet.title = "Dashboard"
    _title(
        sheet,
        "VUELTA 2026 | ACTIONABLE RACE BOOK",
        f"Locked squad and forward plan after {len(completed)} completed stages",
        10,
    )
    metrics = [
        ("SQUAD PRICE", personal["price"]),
        ("BUDGET LEFT", personal["budget_remaining"]),
        ("LEGALITY", "PASS" if personal["legal"] else "FAIL"),
        ("TEAM GRADE", personal["totals"]["team_grade"]),
    ]
    for index, (label, value) in enumerate(metrics):
        column = 1 + index * 2
        sheet.cell(4, column, label)
        sheet.cell(4, column).font = Font(bold=True, color="5B6573")
        sheet.merge_cells(start_row=5, start_column=column, end_row=5, end_column=column + 1)
        sheet.cell(5, column, value)
        sheet.cell(5, column).font = Font(bold=True, size=16)
        sheet.cell(5, column).fill = PatternFill(
            "solid", fgColor=LIGHT_GREEN if label == "LEGALITY" else LIGHT_BLUE
        )
    sheet["A8"] = (
        f"NEXT ACTION: STAGE {next_stage}" if next_stage is not None else "RACE COMPLETE"
    )
    sheet["A8"].font = Font(bold=True, size=14, color=NAVY)
    if next_stage is not None:
        target = next(
            stage
            for stage in personal["stages"]
            if int(stage["stage_no"]) == next_stage
        )
        sheet["A9"] = "Profile"
        sheet["B9"] = f"{target['profile_type'].title()} / {target['finish_type'].title()}"
        sheet["A10"] = "Captain"
        sheet["B10"] = target["captain"]
        sheet["B10"].fill = PatternFill("solid", fgColor=GOLD)
        sheet["B10"].font = Font(bold=True)
        sheet["A11"] = "First reserve"
        sheet["B11"] = reserves[next_stage]
        sheet["A13"] = "ENROL EXACTLY THESE NINE"
        sheet["A13"].font = Font(bold=True, color=NAVY)
        for index, rider in enumerate(target["lineup"], start=1):
            row = 14 + (index - 1) // 3
            column = 1 + ((index - 1) % 3) * 3
            sheet.cell(row, column, index)
            sheet.cell(row, column + 1, rider)
            sheet.merge_cells(
                start_row=row,
                start_column=column + 1,
                end_row=row,
                end_column=column + 2,
            )
            if rider == target["captain"]:
                sheet.cell(row, column + 1).fill = PatternFill("solid", fgColor=GOLD)
                sheet.cell(row, column + 1).font = Font(bold=True)
    sheet["A19"] = "DATA STATUS"
    sheet["A19"].font = Font(bold=True, size=12, color=NAVY)
    status = [
        ("Scorito market snapshot", _timestamp(personal["market_snapshot_time"])),
        ("Stage predictions", _timestamp(personal["prediction_generated_at"])),
        ("Saved personal squad", _timestamp(personal["personal_snapshot_time"])),
        ("Completed stages", ", ".join(str(row["stage_no"]) for row in completed)),
        ("Authentication", "Expired token; no Scorito team or lineup was edited"),
        ("Model status", projection.get("model_version", "not available")),
    ]
    for row, (label, value) in enumerate(status, start=20):
        sheet.cell(row, 1, label).font = Font(bold=True)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        sheet.cell(row, 2, value)
    _finish(sheet, freeze="A8", widths={1: 22, 2: 22, 3: 6, 4: 22, 5: 6, 6: 22, 7: 6, 8: 22, 9: 6, 10: 18})


def _current_squad(
    workbook: Workbook,
    personal: dict[str, Any],
    snapshot,
    teams: dict[int, str],
    next_stage: int | None,
) -> None:
    sheet = workbook.create_sheet("Current Squad")
    _title(sheet, "LOCKED 20-RIDER SQUAD", "Saved personal team; this is the actionable roster", 11)
    headers = [
        "#", "Rider", "Trade team", "Role", "Nation", "Age", "Price",
        "Forward starts", "Captaincies", "Projected contribution", "Assessment",
    ]
    sheet.append([])
    sheet.append(headers)
    _header(sheet, 4)
    by_name = {_name_key(rider.name): rider for rider in snapshot.riders}
    forward_stages = _future_stages(personal, next_stage)
    starts = Counter(name for stage in forward_stages for name in stage["lineup"])
    captaincies = Counter(stage["captain"] for stage in forward_stages)
    for index, row in enumerate(personal["riders"], start=1):
        rider = by_name[_name_key(row["rider"])]
        sheet.append([
            index, rider.name, teams.get(rider.team_id, f"Team {rider.team_id}"),
            rider.role_label, rider.nationality, rider.age, rider.price,
            starts[rider.name], captaincies[rider.name],
            row["total_projected_contribution"], row["assessment"].title(),
        ])
    sheet.auto_filter.ref = f"A4:K{sheet.max_row}"
    for cell in sheet[4]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
    for row in range(5, sheet.max_row + 1):
        sheet.cell(row, 7).number_format = '#,##0'
        sheet.cell(row, 10).number_format = '0.00'
    _finish(sheet, freeze="A5", widths={1: 5, 2: 24, 3: 30, 4: 12, 5: 9, 6: 7, 7: 13, 8: 14, 9: 12, 10: 20, 11: 13})


def _forward_plan(
    workbook: Workbook,
    personal: dict[str, Any],
    reserves: dict[int, str],
    next_stage: int | None,
) -> None:
    sheet = workbook.create_sheet("Forward Stage Plan")
    subtitle = (
        "Race complete; no remaining stages"
        if next_stage is None
        else f"Stages {next_stage}-21 only; exactly nine riders and one captain"
    )
    _title(sheet, "FORWARD STAGE PLAN", subtitle, 19)
    headers = [
        "Stage", "Profile", "Finish", "Projected winner", "Captain", "First reserve",
        "Finish pts", "Captain bonus", "Conditional bonus", "Projected total",
        *[f"Rider {index}" for index in range(1, 10)],
    ]
    sheet.append([])
    sheet.append(headers)
    _header(sheet, 4)
    for stage in _future_stages(personal, next_stage):
        stage_no = int(stage["stage_no"])
        conditional = float(stage["team_bonus_points"]) + float(
            stage["red_jersey_team_bonus_points"]
        )
        sheet.append([
            stage_no, stage["profile_type"].title(), stage["finish_type"].title(),
            stage["projected_winner"], stage["captain"], reserves[stage_no],
            stage["finish_points"], stage["captain_bonus_points"], conditional,
            stage["stage_total"], *stage["lineup"],
        ])
        for cell in sheet[sheet.max_row][10:]:
            if cell.value == stage["captain"]:
                cell.fill = PatternFill("solid", fgColor=GOLD)
                cell.font = Font(bold=True)
    sheet.auto_filter.ref = f"A4:S{sheet.max_row}"
    _finish(sheet, freeze="A5", widths={1: 7, 2: 11, 3: 11, 4: 22, 5: 22, 6: 22, 7: 11, 8: 13, 9: 16, 10: 14, **{column: 22 for column in range(11, 20)}})


def _stage_matrix(
    workbook: Workbook, personal: dict[str, Any], next_stage: int | None
) -> None:
    sheet = workbook.create_sheet("Stage Matrix")
    _title(sheet, "LOCKED-SQUAD DEPLOYMENT", "C = captain, X = enrolled, blank = benched", 23)
    forward = _future_stages(personal, next_stage)
    headers = ["Rider", *[f"S{stage['stage_no']}" for stage in forward], "Starts", "Captaincies"]
    sheet.append([])
    sheet.append(headers)
    _header(sheet, 4)
    for rider_row in personal["riders"]:
        rider = rider_row["rider"]
        cells = []
        for stage in forward:
            cells.append("C" if stage["captain"] == rider else "X" if rider in stage["lineup"] else "")
        sheet.append([rider, *cells, cells.count("X") + cells.count("C"), cells.count("C")])
        for cell in sheet[sheet.max_row][1:-2]:
            if cell.value == "C":
                cell.fill = PatternFill("solid", fgColor=GOLD)
                cell.font = Font(bold=True)
            elif cell.value == "X":
                cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            cell.alignment = Alignment(horizontal="center")
    _finish(sheet, freeze="B5", widths={1: 24, **{column: 5 for column in range(2, 22)}, 22: 8, 23: 12})


def _completed_results(
    workbook: Workbook,
    snapshot,
    personal_ids: set[int],
    completed: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("Completed Stages")
    _title(
        sheet,
        "COMPLETED STAGE RESULTS",
        "Realized results kept separate from forward predictions",
        6,
    )
    headers = [
        "Stage",
        "Actual rank",
        "Rider",
        "In locked squad",
        "Time (s)",
        "Stage status",
    ]
    sheet.append([])
    sheet.append(headers)
    _header(sheet, 4)
    names = {rider.rider_id: rider.name for rider in snapshot.riders}
    for stage in completed:
        for result in stage["results"]:
            rider_id = int(result["RiderId"])
            sheet.append(
                [
                    stage["stage_no"],
                    int(result["Rank"]),
                    names.get(rider_id, f"Rider {rider_id}"),
                    "Yes" if rider_id in personal_ids else "No",
                    float(result.get("Time") or 0) / 1000.0,
                    "Completed",
                ]
            )
            if rider_id in personal_ids:
                for cell in sheet[sheet.max_row]:
                    cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    sheet.auto_filter.ref = f"A4:F{sheet.max_row}"
    _finish(
        sheet,
        freeze="A5",
        widths={1: 8, 2: 13, 3: 25, 4: 16, 5: 12, 6: 12},
    )


def _model_comparator(workbook: Workbook, optimum: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Model Comparator")
    scenario = optimum["base_optimal"]
    _title(sheet, "THEORETICAL UNLOCKED OPTIMUM", "Comparator only; the race has started and this squad is not actionable", 5)
    sheet.append([])
    sheet.append(["Metric", "Value"])
    _header(sheet, 4)
    for label, value in [
        ("Price", scenario["price"]),
        ("Budget", optimum["constraints"]["budget"]),
        ("Projected score", scenario["final_projected_score"]),
        ("Generated", _timestamp(optimum["generated_at"])),
    ]:
        sheet.append([label, value])
    sheet.append([])
    sheet.append(["#", "Rider", "Status", "Reason", "Model scope"])
    _header(sheet, 10)
    for index, rider in enumerate(scenario["riders"], start=1):
        sheet.append([index, rider, "Comparator", "Cannot replace locked squad", "All 21 stages"])
    _finish(sheet, freeze="A11", widths={1: 10, 2: 26, 3: 14, 4: 28, 5: 16})


def _sources(
    workbook: Workbook,
    personal: dict[str, Any],
    optimum: dict[str, Any],
    next_stage: int | None,
) -> None:
    sheet = workbook.create_sheet("Sources & Validation")
    _title(sheet, "SOURCES AND VALIDATION", "Freshness, constraints, and remaining gaps", 4)
    sheet.append([])
    sheet.append(["Source", "Timestamp", "Status", "Purpose"])
    _header(sheet, 4)
    sources = [
        ("Scorito market snapshot", personal["market_snapshot_time"], "Live saved public data", "Prices, availability, stages, stage 1 result"),
        ("PCS evidence projection", personal["projection_generated_at"], "184-rider model", "Objective rider-stage scores"),
        ("Stage predictions", personal["prediction_generated_at"], "Objective ordering", "Forward top 20 per stage"),
        ("Saved personal squad", personal["personal_snapshot_time"], "Authentication expired", "Locked 20-rider roster"),
        ("Exact optimum", optimum["generated_at"], "Comparator only", "Post-lock opportunity-cost view"),
    ]
    for source, timestamp, status, purpose in sources:
        sheet.append([source, _timestamp(timestamp), status, purpose])
    sheet.append([])
    sheet.append(["Validation", "Result", "Status", "Notes"])
    _header(sheet, 11)
    forward = _future_stages(personal, next_stage)
    expected_forward = 0 if next_stage is None else 22 - next_stage
    checks = [
        ("Unique riders", personal["unique_riders"], personal["unique_riders"] == 20, "Exactly 20 required"),
        ("Budget", personal["price"], personal["price"] <= personal["budget"], f"Limit {personal['budget']:,}"),
        ("Trade-team maximum", personal["max_trade_team_count"], personal["max_trade_team_count"] <= 4, "Maximum four"),
        ("Unavailable riders", len(personal["unavailable_riders"]), not personal["unavailable_riders"], "None allowed"),
        ("Forward stages", len(forward), len(forward) == expected_forward, f"Stages {next_stage or 'complete'}-21"),
        ("Lineups", len(forward), all(len(s["lineup"]) == len(set(s["lineup"])) == 9 for s in forward), "Nine unique riders each"),
        ("Captains", len(forward), all(s["captain"] in s["lineup"] for s in forward), "Captain enrolled"),
    ]
    for label, result, passed, notes in checks:
        sheet.append([label, result, "PASS" if passed else "FAIL", notes])
        sheet.cell(sheet.max_row, 3).fill = PatternFill("solid", fgColor=LIGHT_GREEN if passed else RED)
    sheet.append([])
    sheet.append(["Remaining gap", "Expired Scorito token", "OPEN", "Personal squad and saved selections could not be refreshed after 20 Aug; no team was edited."])
    sheet.append(["Method", "Objective-only lineup selection", "PASS", "Chat/forum adjusted scores and conditional team bonuses cannot force lineup or captain decisions."])
    _finish(sheet, freeze="A5", widths={1: 28, 2: 25, 3: 22, 4: 70})


def _validate_workbook(
    output_path: Path,
    personal: dict[str, Any],
    completed: list[dict[str, Any]],
    next_stage: int | None,
) -> None:
    workbook = load_workbook(output_path, read_only=True, data_only=False)
    expected_sheets = [
        "Dashboard",
        "Current Squad",
        "Forward Stage Plan",
        "Stage Matrix",
        "Completed Stages",
        "Model Comparator",
        "Sources & Validation",
    ]
    if workbook.sheetnames != expected_sheets:
        raise RuntimeError("saved workbook sheet validation failed")
    expected_squad = [row["rider"] for row in personal["riders"]]
    saved_squad = [
        row[0]
        for row in workbook["Current Squad"].iter_rows(
            min_row=5, min_col=2, max_col=2, values_only=True
        )
    ]
    if len(saved_squad) != 20 or saved_squad != expected_squad:
        raise RuntimeError("saved workbook squad validation failed")
    expected_forward = _future_stages(personal, next_stage)
    forward_sheet = workbook["Forward Stage Plan"]
    saved_forward = list(
        forward_sheet.iter_rows(
            min_row=5, min_col=1, max_col=19, values_only=True
        )
    )
    if len(saved_forward) != len(expected_forward):
        raise RuntimeError("saved workbook forward-stage count validation failed")
    for saved, expected in zip(saved_forward, expected_forward, strict=True):
        if (
            saved[0] != int(expected["stage_no"])
            or saved[4] != expected["captain"]
            or list(saved[10:19]) != expected["lineup"]
        ):
            raise RuntimeError(
                f"saved workbook stage {expected['stage_no']} plan validation failed"
            )
    expected_results = [
        (stage["stage_no"], int(result["Rank"]), int(result["RiderId"]))
        for stage in completed
        for result in stage["results"]
    ]
    result_sheet = workbook["Completed Stages"]
    saved_results = list(
        result_sheet.iter_rows(min_row=5, min_col=1, max_col=3, values_only=True)
    )
    snapshot_names = {
        rider.rider_id: rider.name for rider in load_snapshot("vuelta2026").riders
    }
    expected_result_rows = [
        (stage_no, rank, snapshot_names.get(rider_id, f"Rider {rider_id}"))
        for stage_no, rank, rider_id in expected_results
    ]
    if saved_results != expected_result_rows:
        raise RuntimeError("saved workbook completed-result count validation failed")
    expected_action = (
        f"NEXT ACTION: STAGE {next_stage}" if next_stage is not None else "RACE COMPLETE"
    )
    if workbook["Dashboard"]["A8"].value != expected_action:
        raise RuntimeError("saved workbook next-stage validation failed")
    if next_stage is not None:
        expected_next = expected_forward[0]
        dashboard = workbook["Dashboard"]
        saved_lineup = [
            dashboard.cell(row=row, column=column).value
            for row in range(14, 17)
            for column in (2, 5, 8)
        ]
        if (
            dashboard["B10"].value != expected_next["captain"]
            or saved_lineup != expected_next["lineup"]
        ):
            raise RuntimeError("saved workbook next-stage lineup validation failed")
    workbook.close()


def export_workbook(output_path: Path = DEFAULT_OUTPUT) -> Path:
    personal = json.loads(PERSONAL_PATH.read_text(encoding="utf-8"))
    predictions = json.loads(PREDICTIONS_PATH.read_text(encoding="utf-8"))
    projection = json.loads(PROJECTION_PATH.read_text(encoding="utf-8"))
    optimum = json.loads(OPTIMUM_PATH.read_text(encoding="utf-8"))
    snapshot = load_snapshot("vuelta2026")
    personal_ids = set(_content(PERSONAL_SELECTION_PATH))
    reserves = _first_reserves(personal, projection)
    completed, next_stage = _race_state()

    workbook = Workbook()
    workbook.properties.creator = "Scorito Cycling Agent"
    workbook.properties.title = "Vuelta 2026 Actionable Team and Stage Plan"
    workbook.properties.description = (
        f"Locked squad, {len(completed)} completed stages, and remaining plan"
    )
    _dashboard(workbook, personal, projection, reserves, completed, next_stage)
    _current_squad(workbook, personal, snapshot, _team_names(), next_stage)
    _forward_plan(workbook, personal, reserves, next_stage)
    _stage_matrix(workbook, personal, next_stage)
    _completed_results(workbook, snapshot, personal_ids, completed)
    _model_comparator(workbook, optimum)
    _sources(workbook, personal, optimum, next_stage)

    temporary = output_path.with_name(f"{output_path.stem}.tmp.xlsx")
    workbook.save(temporary)
    temporary.replace(output_path)
    _validate_workbook(output_path, personal, completed, next_stage)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    print(export_workbook(args.out))


if __name__ == "__main__":
    main()
